# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import re
import copy
import xlrd
try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell

# wb = load_workbook(u'C:/D4/duan_migrate/tonkho/file_import/Mẫu BBBG 2018.xlsx')
# ws = wb.active
# 
# 
# ws.insert_rows(ws,16, 5)
# wb.save(u'C:/D4/test_folder/Mẫu BBBG 2018_new3.xlsx',)
# print 'done'


from xlutils.filter import process,XLRDReader,XLWTWriter
import xlrd, xlwt

# Demonstration of copy2 patch for xlutils 1.4.1

# Context:
# xlutils.copy.copy(xlrd_workbook) -> xlwt_workbook
# copy2(xlrd_workbook) -> (xlwt_workbook, style_list)
# style_list is a conversion of xlrd_workbook.xf_list to xlwt-compatible styles

# Step 1: Create an input file for the demo
def create_input_file():
    wtbook = xlwt.Workbook()
    wtsheet = wtbook.add_sheet(u'First')
    colours = 'white black red green blue pink turquoise yellow'.split()
    fancy_styles = [xlwt.easyxf(
        'font: name Times New Roman, italic on;'
        'pattern: pattern solid, fore_colour %s;'
         % colour) for colour in colours]
    for rowx in xrange(8):
        wtsheet.write(rowx, 0, rowx)
        wtsheet.write(rowx, 1, colours[rowx], fancy_styles[rowx])
    wtbook.save('demo_copy2_in.xls')
# Patch: add this function to the end of xlutils/copy.py
def copy2(wb):
    w = XLWTWriter()
    process(
        XLRDReader(wb,'unknown.xls'),
        w
        )
    return w.output[0][1], w.style_list

# def update_content():
#     rdbook = xlrd.open_workbook(u'C:/D4/test_folder/Mẫu BBBG 2018.xls',formatting_info=True )#formatting_info=True
#     sheetx = 0
#     rdsheet = rdbook.sheet_by_index(sheetx)
#     wtbook, style_list = copy2(rdbook)
#     wtsheet = wtbook.get_sheet(sheetx)
#     fixups = [(0, 0, 'MAGENTA'), (1, 1, 'CYAN')]
#     
#     t = xlwt.easyxf(
#         'font: name Times New Roman, italic on;'
#         'pattern: pattern solid'
#         )
#     print t
#      
#      
#      
#     for rowx, colx, value in fixups:
#         
#         xf_index = rdsheet.cell_xf_index(rowx, colx)
#         print ('xf_index',xf_index)
#         print ('style_list[xf_index]',style_list[xf_index])
#         wtsheet.write(rowx, colx, value, str(style_list[xf_index]))
# #     raise ValueError('dfkldjlf')
#     wtbook.save(u'C:/D4/test_folder/Mẫu BBBG 2018hehe.xls')

# def create_xl ():
#     w = XLWTWriter()
def create_input_file2():
    wtbook = XLWTWriter()
    wtsheet = wtbook.add_sheet(u'First')
    colours = 'white black red green blue pink turquoise yellow'.split()
    fancy_styles = [xlwt.easyxf(
        'font: name Times New Roman, italic on;'
        'pattern: pattern solid, fore_colour %s;'
         % colour) for colour in colours]
    for rowx in xrange(8):
        wtsheet.write(rowx, 0, rowx)
        wtsheet.write(rowx, 1, colours[rowx], fancy_styles[rowx])
    wtbook.save(u'C:/D4/test_folder/Mẫu BBBG 2018_nana.xls')  


def write_merge_cell(row,col,merge_tuple_list):
    for crange in merge_tuple_list:
        rlo, rhi, clo, chi = crange
        if row>=rlo and row < rhi and col >=clo and col < chi:
            row = rlo
            col = clo
            return crange
    return None


def update_content():
    rdbook= xlrd.open_workbook(u'C:/D4/test_folder/Mẫu BBBG 2018.xls',formatting_info=True )#formatting_info=True
    sheetx = 0
    rdsheet = rdbook.sheet_by_index(sheetx)
#     print rdbook.style_list
    print ('rdsheet',rdsheet)
    wtbook_c, style_list = copy2(rdbook)
    wtsheet_c = wtbook_c.get_sheet(sheetx)
    wtbook = xlwt.Workbook()
    wtsheet = wtbook.add_sheet(u'First')
    for row in range(1,rdsheet.ncols):
        wtsheet.row(row).height_mismatch = True
        wtsheet.row(row).height = wtsheet_c.row(row).height 
    for col in range(1,rdsheet.ncols):
        wtsheet.col(col).width =  wtsheet_c.col(col).width
    merge_tuple_list =  rdsheet.merged_cells
    print ('merge_tuple_list',merge_tuple_list)
#     wtsheet = sss.get_sheet(sheetx)
    fixups = [(0, 0), (1, 0),(0,3),(1,3),
                    (3,0),(5,0),(6,0),
                     (9,0),(9,2)
              ]
#     fixups = [(0, 0)]
    t = xlwt.easyxf(
        'font: name Times New Roman, italic on;'
        'pattern: pattern solid'
        )
    print t
     
     
     
    for rowx, colx in fixups:
        xf_index = rdsheet.cell_xf_index(rowx, colx)
        val = rdsheet.cell_value(rowx,colx)
        crange = write_merge_cell(rowx, colx, merge_tuple_list)
        print ('crange',crange)
        if crange ==None:
            wtsheet.write(rowx, colx, val, style_list[xf_index])
        else:
            wtsheet.write_merge(crange[0],crange[1]-1,crange[2],crange[3]-1, val, style_list[xf_index])
    wtbook.save(u'C:/D4/test_folder/Mẫu BBBG 2018hehe.xls')
    
    

update_content()
# update_content()

